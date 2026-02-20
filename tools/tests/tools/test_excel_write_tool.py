from pathlib import Path
from unittest.mock import patch

import pytest
from fastmcp import FastMCP

from aden_tools.tools.excel_write_tool.excel_write_tool import register_tools

TEST_WORKSPACE_ID = "test-workspace"
TEST_AGENT_ID = "test-agent"
TEST_SESSION_ID = "test-session"


@pytest.fixture
def excel_tools(mcp: FastMCP, tmp_path: Path):
    with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
        register_tools(mcp)
        yield {
            "excel_write": mcp._tool_manager._tools["excel_write"].fn,
            "tmp_path": tmp_path,
        }


def test_excel_write_multisheet_and_format(excel_tools):
    fn = excel_tools["excel_write"]

    wb = {
        "sheets": [
            {
                "name": "Summary",
                "columns": ["Ticker", "Return", "Drawdown"],
                "rows": [["AAPL", 0.0123, -0.034], ["MSFT", 0.004, -0.021]],
                "freeze_panes": "A2",
                "number_formats": {"Return": "0.00%", "Drawdown": "0.00%"},
            },
            {
                "name": "Raw",
                "columns": ["Day", "AAPL", "MSFT"],
                "rows": [["2026-02-01", 192.1, 411.2]],
                "freeze_panes": "A2",
            },
        ]
    }

    res = fn(
        path="out/weekly_report.xlsx",
        workbook=wb,
        workspace_id=TEST_WORKSPACE_ID,
        agent_id=TEST_AGENT_ID,
        session_id=TEST_SESSION_ID,
    )
    assert res.get("success") is True
    assert res.get("metadata", {}).get("sheets") == 2


    out_abs = (
        Path(excel_tools["tmp_path"])
        / TEST_WORKSPACE_ID
        / TEST_AGENT_ID
        / TEST_SESSION_ID
        / "out"
        / "weekly_report.xlsx"
    )
    assert out_abs.exists()
    assert out_abs.stat().st_size > 0

    # Light read-back checks
    import openpyxl
    wb2 = openpyxl.load_workbook(out_abs)
    assert "Summary" in wb2.sheetnames
    ws = wb2["Summary"]
    assert ws["A1"].value == "Ticker"
    assert ws.freeze_panes == "A2"
    # number format applied
    assert ws["B2"].number_format == "0.00%"


def test_excel_write_embeds_image(tmp_path: Path):
    import openpyxl
    from fastmcp import FastMCP

    from aden_tools.tools.chart_tool.chart_tool import register_tools as register_chart
    from aden_tools.tools.excel_write_tool.excel_write_tool import register_tools as register_xlsx
    from aden_tools.tools.file_system_toolkits.security import get_secure_path

    with patch("aden_tools.tools.file_system_toolkits.security.WORKSPACES_DIR", str(tmp_path)):
        mcp = FastMCP("img-xlsx-test")
        register_chart(mcp)
        register_xlsx(mcp)

        chart_png = mcp._tool_manager._tools["chart_render_png"].fn
        xlsx = mcp._tool_manager._tools["excel_write"].fn

        r = chart_png(
            path="out/chart.png",
            chart={
                "title": "T",
                "x_label": "x",
                "y_label": "y",
                "series": [{"name": "s", "x": [1, 2], "y": [1, 2]}],
            },
            workspace_id="w",
            agent_id="a",
            session_id="s",
        )
        assert r["success"] is True, r

        r2 = xlsx(
            path="out/book.xlsx",
            workbook={
                "sheets": [{
                    "name": "Charts",
                    "columns": ["Chart"],
                    "rows": [["demo"]],
                    "images": [{"path": "out/chart.png", "cell": "A3", "width": 400}],
                }]
            },
            workspace_id="w",
            agent_id="a",
            session_id="s",
        )
        assert r2["success"] is True, r2

        x_abs = get_secure_path("out/book.xlsx", "w", "a", "s")
        wb = openpyxl.load_workbook(x_abs)
        ws = wb["Charts"]
        assert len(getattr(ws, "_images", [])) >= 1


def test_excel_write_update_mode_replaces_sheet(excel_tools):
    fn = excel_tools["excel_write"]

    res1 = fn(
        path="out/update_report.xlsx",
        workbook={
            "sheets": [
                {
                    "name": "Summary",
                    "columns": ["Ticker", "Return"],
                    "rows": [["AAPL", 0.01]],
                }
            ]
        },
        workspace_id=TEST_WORKSPACE_ID,
        agent_id=TEST_AGENT_ID,
        session_id=TEST_SESSION_ID,
        mode="create",
    )
    assert res1.get("success") is True, res1

    res2 = fn(
        path="out/update_report.xlsx",
        workbook={
            "sheets": [
                {
                    "name": "Summary",
                    "columns": ["Ticker", "Return"],
                    "rows": [["AAPL", 0.25]],
                }
            ]
        },
        workspace_id=TEST_WORKSPACE_ID,
        agent_id=TEST_AGENT_ID,
        session_id=TEST_SESSION_ID,
        mode="update",
    )
    assert res2.get("success") is True, res2

    out_abs = (
        Path(excel_tools["tmp_path"])
        / TEST_WORKSPACE_ID
        / TEST_AGENT_ID
        / TEST_SESSION_ID
        / "out"
        / "update_report.xlsx"
    )
    import openpyxl

    wb2 = openpyxl.load_workbook(out_abs)
    ws = wb2["Summary"]
    assert ws["B2"].value == 0.25
