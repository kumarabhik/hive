"""Excel Write Tool - Generate .xlsx workbooks (schema-first)."""
from __future__ import annotations
from aden_tools.tools.office_skills_pack.contracts import ArtifactError, ArtifactResult


import os
from typing import Any, Dict, List, Optional

from fastmcp import FastMCP
from pydantic import BaseModel, Field

from ..file_system_toolkits.security import get_secure_path


class SheetImageSpec(BaseModel):
    path: str = Field(..., min_length=1, description="Image path relative to session sandbox")
    cell: str = Field(default="A1", description="Top-left anchor cell, e.g. C3")
    width: Optional[int] = Field(default=None, ge=1)
    height: Optional[int] = Field(default=None, ge=1)


class SheetSpec(BaseModel):
    name: str = Field(..., min_length=1)
    columns: List[str] = Field(..., min_length=1)
    rows: List[List[Any]] = Field(default_factory=list)

    # "B2" style or "A2" style. Common: "A2" = freeze header row.
    freeze_panes: Optional[str] = Field(default="A2")

    # Map column name -> excel number format string, e.g. {"Return": "0.00%", "Price": "$#,##0.00"}
    number_formats: Dict[str, str] = Field(default_factory=dict)

    # Optional fixed widths: {"Ticker": 12, "Return": 10}
    column_widths: Dict[str, float] = Field(default_factory=dict)
    images: List[SheetImageSpec] = Field(
        default_factory=list,
        description="Images to place in worksheet",
    )


class WorkbookSpec(BaseModel):
    sheets: List[SheetSpec] = Field(..., min_length=1)


def _best_effort_autosize(ws) -> None:
    # Best-effort autosize based on cell string length (cheap + good enough for MVP)
    try:
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        if max_len > 0:
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)


def register_tools(mcp: FastMCP) -> None:
    @mcp.tool()
    def excel_write(
        path: str,
        workbook: dict[str, Any],
        workspace_id: str,
        agent_id: str,
        session_id: str,
    ) -> dict:
        """
        Write an Excel (.xlsx) file from a strict schema into the session sandbox.
        """

        if not path.lower().endswith(".xlsx"):
            return ArtifactResult(
                success=False,
                error=ArtifactError(
                    code="INVALID_SCHEMA",
                    message="path must end with .xlsx",
                ),
            ).model_dump()




        try:
            spec = WorkbookSpec.model_validate(workbook)
        except Exception as e:
            return ArtifactResult(
                success=False,
                error=ArtifactError(
                    code="INVALID_SCHEMA",
                    message="Invalid workbook schema",
                    details=str(e),
                ),
            ).model_dump()


        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:

            return ArtifactResult(
                success=False,
                error=ArtifactError(code="DEP_MISSING", message="openpyxl not installed. Install with: pip install -e '.[excel]'"),
            ).model_dump()


        try:
            out_path = get_secure_path(path, workspace_id, agent_id, session_id)
            os.makedirs(os.path.dirname(out_path), exist_ok=True)

            wb = Workbook()
            # remove default sheet
            if wb.sheetnames:
                wb.remove(wb[wb.sheetnames[0]])

            header_font = Font(bold=True)


            if "sheets" not in workbook:
                return ArtifactResult(
                    success=False,
                    error=ArtifactError(
                        code="INVALID_SCHEMA",
                        message="Invalid workbook structure: 'sheets' key not found",
                    ),
                ).model_dump()


            for sheet in spec.sheets:
                ws = wb.create_sheet(title=sheet.name[:31])  # Excel sheet name limit

                # header
                for j, col in enumerate(sheet.columns, start=1):
                    c = ws.cell(row=1, column=j, value=col)
                    c.font = header_font

                # rows
                for i, row in enumerate(sheet.rows, start=2):
                    for j in range(1, len(sheet.columns) + 1):
                        v = row[j - 1] if j - 1 < len(row) else None
                        ws.cell(row=i, column=j, value=v)

                # freeze panes
                if sheet.freeze_panes:
                    ws.freeze_panes = sheet.freeze_panes

                # apply number formats by column name
                col_to_idx = {name: idx + 1 for idx, name in enumerate(sheet.columns)}
                for col_name, fmt in sheet.number_formats.items():
                    idx = col_to_idx.get(col_name)
                    if not idx:
                        continue
                    for r in range(2, ws.max_row + 1):
                        ws.cell(row=r, column=idx).number_format = fmt

                # explicit column widths
                for col_name, width in sheet.column_widths.items():
                    idx = col_to_idx.get(col_name)
                    if not idx:
                        continue
                    from openpyxl.utils import get_column_letter
                    ws.column_dimensions[get_column_letter(idx)].width = float(width)

                # autosize as fallback
                _best_effort_autosize(ws)

                # images
                if sheet.images:
                    from openpyxl.drawing.image import Image as XLImage

                    for img in sheet.images:
                        ext = os.path.splitext(img.path.lower())[1]
                        if ext not in [".png", ".jpg", ".jpeg"]:
                            return ArtifactResult(
                                success=False,
                                error=ArtifactError(
                                    code="INVALID_PATH",
                                    message=f"unsupported image extension: {ext}",
                                ),
                            ).model_dump()
                        img_abs = get_secure_path(img.path, workspace_id, agent_id, session_id)
                        if os.path.exists(img_abs):
                            xl_img = XLImage(img_abs)
                            if img.width:
                                xl_img.width = int(img.width)
                            ws.add_image(xl_img, img.cell)

            wb.save(out_path)

            
        
            return ArtifactResult(
                success=True,
                output_path=str(out_path),
                metadata={"sheets": len(spec.sheets)},
            ).model_dump()


        except Exception as e:
            return ArtifactResult(
                success=False,
                error=ArtifactError(
                    code="INTERNAL_ERROR",
                    message="Failed to generate Excel workbook",
                    details=str(e),
                ),
            ).model_dump()
